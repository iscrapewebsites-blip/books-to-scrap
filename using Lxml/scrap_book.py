from lxml import html
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def get_data(res, page_no):
     tree = html.fromstring(res.text)
     elem = tree.xpath('//section//ol/li')

     # xpath() always returns a list

     data = []

     for i, e in enumerate(elem):
          dict = {}
          dict['S.No.'] = i+1+page_no*20
          dict['Title'] = e.xpath('.//@title')[0]
          dict['Price'] = e.xpath('.//p[@class="price_color"]')[0].text_content().split('Â')[1]
          dict['Rating'] = e.xpath('.//p[contains(@class, "star-rating")]/@class')[0].split(' ')[1]

          data.append(dict)
     
     return data
          

def data_to_xls(data, isFirst):
     if isFirst==True:
          wb = Workbook()
          ws = wb.active
          header = list(data[0].keys())
          ws.append(header)

          # setting header bold
          for cell in ws['1']:
               cell.font = Font(bold=True)
          
          for row in data:
               ws.append(list(row.values()))
          
          wb.save('books.xlsx')
     else:
          wb = load_workbook('books.xlsx')
          ws = wb.active
          for row in data:
               ws.append(list(row.values()))
          wb.save('books.xlsx')

for page_no in range(1,51):
     url = f'https://books.toscrape.com/catalogue/page-{page_no}.html'
     response = requests.get(url)
     data = get_data(response, page_no-1)
     isFirst = False
     if page_no==1:
          isFirst = True
     data_to_xls(data, isFirst)
     